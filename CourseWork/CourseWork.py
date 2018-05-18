import os
from openpyxl import load_workbook

table = [
#   ('festival', 'espeak', 'sapi', 'cepstral', 'mac', 'x-sampa', 'acapela-uk', 'cmu', 'bbcmicro', 'unicode-ipa','pinyin-approx'),
   # The first entry MUST be the syllable separator:
   ('0', '%', '-', '0', '=', '.', '', '0', '', '.',''),
   ('1', "'", '1', '1', '1', '"', 0, '1', '1', u'\u02c8','4'), # primary stress - ignoring this for acapela-uk
   ('2', ',', '2', '0', '2', '%', 0, '2', '2', u'\u02cc','2'), # secondary stress - ignoring this for acapela-uk
   ('', '', '', '', '', '', 0, '', '', '-',''),
   (0, 0, 0, 0, 0, 0, 0, 0, '', '#',0),
   (0, 0, 0, 0, 0, 0, 0, 0, '', ' ',0),
   (0, 0, 0, 0, 0, 0, 0, 0, '', '_',0),
   (0, 0, 0, 0, 0, 0, 0, 0, '', '?',0),
   (0, 0, 0, 0, 0, 0, 0, 0, '', '!',0),
   (0, 0, 0, 0, 0, 0, 0, 0, '', ',',0),
   ('aa', ['A:', 'A@', 'aa'], 'aa', 'a', 'AA', 'A', 'A:', 'AA', 'AA', u'\u0251','a5'),
   (0, 'A', 0, 0, 0, 0, 0, '2', 0, u'\u2051',0),
   (0, 'A:', 0, 0, 0, ':', 0, '1', 0, u'\u02d0',0),
   (0, 0, 0, 0, 0, 'A:', 0, 'AA', 0, u'\u0251\u02d0',0),
   (0, 0, 0, 0, 0, 'Ar\\', 0, 0, 0, u'\u0251\u0279',0),
   (0, 0, 0, 0, 'aa', 'a:', 0, 0, 0, 'a\u02d0',0),
   ('a', ['a', '&'], 'ae', 'ae', 'AE', '{', '{', 'AE', 'AE', [u'\xe6','a'],'ya5'),
   ('uh', 'V', 'ah', 'ah', 'UX', 'V', 'V', 'AH', 'OH', u'\u028c','e5'),
   ('o', '0', 'ao', 'oa', 'AA', 'Q', 'Q', 'AA', 'O', u'\u0252','yo5'),
   (0, 0, 0, 0, 0, 'A', 'A', 0, 0, u'\u0251',0),
   (0, 0, 0, 0, 0, 'O', 'O', 0, 0, u'\u0254',0),
   ('au', 'aU', 'aw', 'aw', 'AW', 'aU', 'aU', 'AW', 'AW', u'a\u028a','ao5'),
   (0, 0, 0, 0, 0, '{O', '{O', 0, 0, u'\xe6\u0254',0),
   ('@', '@', 'ax', 'ah', 'AX', '@', '@', 'AH', 'AH', u'\u0259','e5'),
   ('@@', '3:', 'er', 'er', 0, '3:', '3:', 'ER', 'ER', u'\u0259\u02d0','e5'),
   ('@', '3', 'ax', 'ah', 0, '@', '@', 'AH', 'AH', u'\u025a','e5'),
   ('@1', 'a2', 0, 0, 0, 0, 0, 0, 0, u'\u0259', 0),
   ('@2', '@', 0, 0, 0, 0, 0, 0, 0, 0, 0),
   ('ai', 'aI', 'ay', 'ay', 'AY', 'aI', 'aI', 'AY', 'IY', u'a\u026a','ai5'),
   (0, 0, 0, 0, 0, 'Ae', 'A e', 0, 0, u'\u0251e',0),
   ('b', 'b', 'b', 'b', 'b', 'b', 'b', 'B', 'B', 'b','bu0'),
   ('ch', 'tS', 'ch', 'ch', 'C', 'tS', 't S', 'CH', 'CH', [u't\u0283', u'\u02a7'],'che0'),
   ('d', 'd', 'd', 'd', 'd', 'd', 'd', 'D', 'D', 'd','de0'),
   ('dh', 'D', 'dh', 'dh', 'D', 'D', 'D', 'DH', 'DH', u'\xf0','ze0'),
   ('e', 'E', 'eh', 'eh', 'EH', 'E', 'e', 'EH', 'EH', u'\u025b','ye5'),
   (0, 0, 'ey', 0, 0, 'e', 0, 0, 0, 'e',0),
   ('@@', '3:', 'er', 'er', 'AX', '3:', '3:', 'ER', 'ER', [u'\u025d', u'\u025c\u02d0'],'e5'),
   ('e@', 'e@', 'eh r', 'e@', 'EH r', 'E@', 'e @', 0, 'AI', u'\u025b\u0259','ye5'),
   (0, 0, 0, 0, 0, 'Er\\', 'e r', 0, 0, u'\u025b\u0279',0),
   (0, 0, 0, 0, 0, 'e:', 'e :', 0, 0, u'e\u02d0',0),
   (0, 0, 0, 0, 0, 'E:', 0, 0, 0, u'\u025b\u02d0',0),
   (0, 0, 0, 0, 0, 'e@', 'e @', 0, 0, u'e\u0259',0),
   ('ei', 'eI', 'ey', 'ey', 'EY', 'eI', 'eI', 'EY', 'AY', u'e\u026a','ei5'),
   (0, 0, 0, 0, 0, '{I', '{I', 0, 0, u'\xe6\u026a',0),
   ('f', 'f', 'f', 'f', 'f', 'f', 'f', 'F', 'F', 'f','fu0'),
   ('g', 'g', 'g', 'g', 'g', 'g', 'g', 'G', 'G', [u'\u0261', 'g'],'ge0'),
   ('h', 'h', 'h', 'h', 'h', 'h', 'h', 'HH', '/H', 'h','he0'),
   ('i', 'I', 'ih', 'ih', 'IH', 'I', 'I', 'IH', 'IH', u'\u026a','yi5'),
   (0, 0, 0, 0, 0, '1', '1', 0, 0, u'\u0268',0),
   (0, ['I', 'I2'], 0, 0, 'IX', 'I', 'I', 0, 'IX', u'\u026a',0),
   ('i@', 'i@', 'iy ah', 'i ah', 'IY UX', 'I@', 'I@', 'EY AH', 'IXAH', u'\u026a\u0259','yi3re5'),
   (0, 0, 0, 0, 0, 'Ir\\', 'I r', 0, 0, u'\u026a\u0279',0),
   ('ii', ['i:','i'], 'iy', 'i', 'IY', 'i', 'i', 'IY', 'EE', 'i','yi5'),
   (0, 0, 0, 0, 0, 'i:', 'i:', 0, 0, u'i\u02d0',0),
   ('jh', 'dZ', 'jh', 'jh', 'J', 'dZ', 'dZ', 'JH', 'J', [u'd\u0292', u'\u02a4'],'zhe0'),
   ('k', 'k', 'k', 'k', 'k', 'k', 'k', 'K', 'K', 'k','ke0'),
   (0, 'x', 0, 0, 0, 'x', 'x', 0, 0, 'x',0),
   ('l', ['l', 'L'], 'l', 'l', 'l', 'l', 'l', 'L', 'L', ['l', u'd\u026b'],'le0'),
   ('m', 'm', 'm', 'm', 'm', 'm', 'm', 'M', 'M', 'm','me0'),
   ('n', 'n', 'n', 'n', 'n', 'n', 'n', 'N', 'N', 'n','ne0'),
   ('ng', 'N', 'ng', 'ng', 'N', 'N', 'N', 'NG', 'NX', u'\u014b','eng0'),
   ('ou', 'oU', 'ow', 'ow', 'OW', '@U', '@U', 'OW', 'OW', [u'\u0259\u028a', 'o'],'ou5'),
   (0, 0, 0, 0, 0, 'oU', 'o U', 0, 0, u'o\u028a',0),
   (0, 0, 0, 0, 0, '@}', '@ }', 0, 0, u'\u0259\u0289',0),
   ('oi', 'OI', 'oy', 'oy', 'OY', 'OI', 'OI', 'OY', 'OY', u'\u0254\u026a','ruo2yi5'),
   (0, 0, 0, 0, 0, 'oI', 'o I', 0, 0, u'o\u026a',0),
   ('p', 'p', 'p', 'p', 'p', 'p', 'p', 'P', 'P', 'p','pu0'),
   ('r', 'r', 'r', 'r', 'r', 'r\\', 'r', 'R', 'R', u'\u0279','re0'),
   (0, 0, 0, 0, 0, 'r', 0, 0, 0, 'r',0),
   ('s', 's', 's', 's', 's', 's', 's', 'S', 'S', 's','se0'),
   ('sh', 'S', 'sh', 'sh', 'S', 'S', 'S', 'SH', 'SH', u'\u0283','she0'),
   ('t', 't', 't', 't', 't', 't', 't', 'T', 'T', ['t', u'\u027e'],'te0'),
   ('th', 'T', 'th', 'th', 'T', 'T', 'T', 'TH', 'TH', u'\u03b8','zhe0'),
   ('u@', 'U@', 'uh', 'uh', 'UH', 'U@', 'U@', 'UH', 'UH', u'\u028a\u0259','wu5'),
   (0, 0, 0, 0, 0, 'Ur\\', 'U r', 0, 0, u'\u028a\u0279',0),
   ('u', 'U', 0, 0, 0, 'U', 'U', 0, '/U', u'\u028a',0),
   ('uu', 'u:', 'uw', 'uw', 'UW', '}:', 'u:', 'UW', ['UW','UX'], u'\u0289\u02d0','yu5'),
   (0, 0, 0, 0, 0, 'u:', 0, 0, 0, [u'u\u02d0', 'u'],0),
   ('oo', 'O:', 'AO', 'ao', 'AO', 'O:', 'O:', 'AO', 'AO', u'\u0254\u02d0','huo5'),
   (0, 0, 0, 0, 0, 'O', 'O', 0, 0, u'\u0254',0),
   (0, 0, 0, 0, 0, 'o:', 'O:', 0, 0, u'o\u02d0',0),
   (0, ['O@', 'o@', 'O'], 0, 0, 0, 'O:', 0, 0, 0, u'\u0254\u02d0',0),
   ('v', 'v', 'v', 'v', 'v', 'v', 'v', 'V', 'V', 'v','fu0'),
   ('w', 'w', 'w', 'w', 'w', 'w', 'w', 'W', 'W', 'w','wu0'),
   (0, 0, 0, 0, 0, 'W', 0, 0, 0, u'\u028d',0),
   ('y', 'j', 'y', 'j', 'y', 'j', 'j', 'Y', 'Y', 'j','yu0'),
   ('z', 'z', 'z', 'z', 'z', 'z', 'z', 'Z', 'Z', 'z','ze0'),
   ('zh', 'Z', 'zh', 'zh', 'Z', 'Z', 'Z', 'ZH', 'ZH', u'\u0292','zhe0'),
   ('@', '@', '@', 'ah', 'AX', '@', '@', '@', 'AH', u'\u0259','wu5'),
   (0, 0, 'ax', '@', 0, 0, 0, 0, 0, 0,0),
   (0, 0, 0, 'ah', '@', 0, 0, 0, 0, 0,0),
   (0, 0, 0, 0, 'AX', 0, 0, 0, '@', '@',0),
]

##################################################################

#Build CMU - x-Sampa table
cmu_sampa_table = list()
for line in table:
    if line[7] != 0 and line[5] != 0:
        cmu_sampa_table.append((line[7], line[5]))

#Build UTF8 - x-Sampa table
utf8_sampa_table = list()
for line in table:
    if line[9] != 0 and line[5] != 0:
        if isinstance(line[9], (list)):
            for variant in line[9]:
                utf8_sampa_table.append((variant, line[5]))
        else:
            utf8_sampa_table.append((line[9], line[5]))

##################################################################

#Build X-SAMPA - CZloid dictionary
wb = load_workbook('D:\VCCV.xlsx')
trans_table = wb.active

sampa_cz_table = list()
#sampa_cz_table.append(('czloid', 'x-sampa'))

#Get vowels list
vowel_set = set()
for vowel_line in trans_table['A3':'B40']:
    cz = vowel_line[0].value
    sampa = vowel_line[1].value
    if isinstance(cz, (int)):
        cz = str(cz)
    if isinstance(sampa, (int)):
        sampa = str(sampa)
    vowel_set.add(cz)
    sampa_cz_table.append((sampa, cz))

#Get CV consonants list
consonant_set = set()
for cv_cons_line in trans_table['A84':'B115']:
    cz = cv_cons_line[0].value
    sampa = cv_cons_line[1].value
    if isinstance(cz, (int)):
        cz = str(cz)
    if isinstance(sampa, (int)):
        sampa = str(sampa)
    consonant_set.add(cz)
    sampa_cz_table.append((sampa, cz))

#Get CCV consonants list
#ccv_cons_set = set()
for ccv_cons_line in trans_table['A43':'B81']:
    cz = ccv_cons_line[0].value
    sampa = ccv_cons_line[1].value
    if isinstance(cz, (int)):
        cz = str(cz)
    if isinstance(sampa, (int)):
        sampa = str(sampa)
    #ccv_cons_set.add(cz)
    #consonant_set.add(cz)
    sampa_cz_table.append((sampa, cz))

#Get VCC consonants list
#vcc_cons_set = set()
for vcc_cons_line in trans_table['A118':'B179']:
    cz = vcc_cons_line[0].value
    sampa = vcc_cons_line[1].value
    if isinstance(cz, (int)):
        cz = str(cz)
    if isinstance(sampa, (int)):
        sampa = str(sampa)
    #vcc_cons_set.add(cz)
    #consonant_set.add(cz)
    sampa_cz_table.append((sampa, cz))

print(sampa_cz_table, end ='\n\n')
sampa_cz_dict = dict(sampa_cz_table)

#Build CMU - CZloid table
cmu_cz_table = list()
cmu_list = list()
sampa_1_list = list()
sampa_2_list = list()
cz_list = list()

for line in cmu_sampa_table:
    cmu_list.append(line[0])
    sampa_1_list.append(line[1])
for line in sampa_cz_table:
    sampa_2_list.append(line[1])
    cz_list.append(line[0])

#print(len(cmu_list), len(sampa_1_list), len(sampa_2_list), len(cz_list), '|', len(vowel_set), '+', len(consonant_set))

#From CMU to CZloid
#cmu_index = 0
#for symbol in sampa_1_list:
#    try:
#        cz_index = 0
#        while True:
#            cz_index = sampa_2_list.index(symbol, cz_index)
#            cmu_cz_table.append((cmu_list[cmu_index], cz_list[cz_index]))
#            cz_index += 1
#    except ValueError:
#        pass
#    cmu_index += 1

#############################
#----- НАЧИНАТЬ ОТСЮДА -----#
#############################

#Build IPA - X-SAMPA dictionary
wb = load_workbook('D:\IPA-X-SAMPA.xlsx')
trans_table = wb.active

ipa_sampa_table = list()
ipa_list = list()
sampa_list = list()
for test_line in trans_table['A2':'B164']:
    ipa = test_line[0].value
    sampa = test_line[1].value
    if (ipa != None and sampa != None):
        if isinstance(ipa, (int)):
            ipa = str(ipa)
        if isinstance(sampa, (int)):
            sampa = str(sampa)
        ipa_sampa_table.append((ipa, sampa))
        ipa_list.append(ipa)
        sampa_list.append(sampa)

print(ipa_sampa_table, end ='\n\n')
ipa_sapma_dict = dict(ipa_sampa_table)

#Build ARPABET - IPA dictionary
wb = load_workbook('D:\ARPABET-IPA.xlsx')
trans_table = wb.active

arpa_ipa_table = list()
ipa_list = list()
arpa_list = list()
for test_line in trans_table['A2':'B63']:
    ipa = test_line[0].value
    arpa = test_line[1].value
    if (ipa != None and arpa != None):
        if isinstance(ipa, (int)):
            ipa = str(ipa)
        if isinstance(arpa, (int)):
            arpa = str(arpa)
        arpa_ipa_table.append((arpa, ipa))
        arpa_list.append(arpa)
        ipa_list.append(ipa)

print(arpa_ipa_table, end ='\n\n')
arpa_ipa_dict = dict(arpa_ipa_table)

for i in set(ipa_list):
    k = ipa_list.count(i)
    if k > 1:
        print(i, k, ipa_list.index(i))
    
#with open(os.path.dirname(__file__) + 'table1.txt', 'w') as f:

#print('[')
#for new_line in new_table:
#    print('\t', new_line, end = ',\n')
#print(']')
